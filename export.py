import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

COLUMNS = [
    ("Ссылка на товар", "url"),
    ("Артикул", "article"),
    ("Название", "name"),
    ("Цена", "price"),
    ("Описание", "description"),
    ("Ссылки на изображения через запятую", "image_urls"),
    ("Все характеристики с сохранением их структуры", "characteristics"),
    ("Название селлера", "seller_name"),
    ("Ссылка на селлера", "seller_url"),
    ("Размеры товара через запятую", "sizes"),
    ("Остатки по товару (число)", "stock"),
    ("Рейтинг", "rating"),
    ("Количество отзывов", "feedbacks"),
]

_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="6B2FA0", end_color="6B2FA0", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _create_workbook(products: list[dict[str, Any]], sheet_title: str) -> Workbook:
    """
    Создать Workbook с данными товаров.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    ws.freeze_panes = "A2"

    for row_idx, product in enumerate(products, start=2):
        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            value = product.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = _CELL_ALIGNMENT
            cell.border = _THIN_BORDER

    _auto_fit_columns(ws)

    return wb


def _auto_fit_columns(ws) -> None:
    """Автоматически подогнать ширину колонок по содержимому."""
    for col_idx in range(1, len(COLUMNS) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted_width


def export_full_catalog(
    products: list[dict[str, Any]],
    filename: str = "wildberries_catalog.xlsx",
) -> str:
    """
    Экспортировать полный каталог в XLSX.
    """
    wb = _create_workbook(products, sheet_title="Каталог")
    wb.save(filename)
    logger.info("Полный каталог сохранён: %s (%d товаров)", filename, len(products))
    return filename


def export_filtered_catalog(
    products: list[dict[str, Any]],
    filename: str = "wildberries_filtered.xlsx",
    min_rating: float = 4.5,
    max_price: float = 10_000,
    country: str = "Россия",
) -> str:
    """
    Экспортировать фильтрованный каталог в XLSX.
    """
    filtered = [
        p for p in products
        if p.get("rating", 0) >= min_rating
        and 0 < p.get("price", 0) <= max_price
        and _matches_country(p, country)
    ]

    wb = _create_workbook(filtered, sheet_title="Фильтр")
    wb.save(filename)
    logger.info(
        "Фильтрованный каталог сохранён: %s (%d из %d товаров)",
        filename, len(filtered), len(products),
    )
    return filename


def _matches_country(product: dict[str, Any], target_country: str) -> bool:
    """Проверить совпадение страны производства."""
    product_country = product.get("country", "")
    if not product_country:
        return False
    return product_country.strip().lower() == target_country.strip().lower()
